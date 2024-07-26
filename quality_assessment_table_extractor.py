import pdfplumber
import pandas as pd
import os
import codecs
import shutil
from openpyxl import load_workbook
from typing import List, Dict, Optional, Tuple
from judge_validation import has_quality_assessment
from use_function import get_title, merge_tables, read_caption
from ai_tool import judge_quality_assessment, read_citation

def write_to_excel(table: List[List[str]], pmid: str, title: str, citation: str):
    """Write the table and citation to the excel file which is located under the data directory"""
    if len(table) == 1:
        if citation != '0':
            explanation_row = [citation] + [''] * (len(table[0][0]) - 1)
            table[0].append(explanation_row)

        df = pd.DataFrame(table[0][1:], columns=table[0][0])
        name = f"{pmid}_{title[:8]}"
        if not os.path.exists('data/output.xlsx'):
            writer = pd.ExcelWriter('data/output.xlsx', engine='openpyxl')
        else:
            book = load_workbook("data/output.xlsx")
            writer = pd.ExcelWriter('data/output.xlsx', engine='openpyxl')
            writer.book = book
        
        df.to_excel(writer, sheet_name=name, index=False)
        
        writer.close()
    else:
        if citation != '0':
            explanation_row = [citation] + [''] * (len(table[0][0]) - 1)
            table.append(explanation_row)

        df = pd.DataFrame(table[1:], columns=table[0])
        name = f"{pmid}_{title[:8]}"
        if not os.path.exists('data/output.xlsx'):
            writer = pd.ExcelWriter('data/output.xlsx', engine='openpyxl')
        else:
            book = load_workbook("data/output.xlsx")
            writer = pd.ExcelWriter('data/output.xlsx', engine='openpyxl')
            writer.book = book
        
        df.to_excel(writer, sheet_name=name, index=False)
        
        writer.close()

def is_quality_assessment_table(title: str, table_type: str):
    if title == None:
        return '0'

    input_value = f"{table_type} : {title}"
    result = judge_quality_assessment(input_value)
    return result

def get_word(words: List[Dict[str, str]]) -> str:
    """Get the word in one page"""
    text = ''
    for word in words:
        text = text + word['text'] + ' '
    
    return text

def merge_word(words: List[str]):
    """Merge words in different pages"""
    text = ''
    for word in words:
        text = text + word + '\n'
    
    return text

def extract_quality_assessment_table(pdf_path: str, pmid: str, table_type: str = "Quality Assessment"):
    """
    Extract quality assessment tables from a PDF file.

    Parameters:
        pdf_path (str): The path to the PDF file.
        pmid (str): Article PMID.
        table_type (str): Type of table
    """

    print(f"Reading file: {pmid}")
    with pdfplumber.open(pdf_path) as pdf:
        is_find = 0
        final_table = []
        final_citation = []
        for i, page in enumerate(pdf.pages):
            words = page.extract_words()
            tables = page.extract_tables()
            for table in tables:
                title = get_title(words)
                result = is_quality_assessment_table(title, table_type)
                if result == '1':
                    print(f'{title} is a quality assessment table')

                # The way to deal with table that cover more than one page and find the citation of table
                if is_find == 0:
                    if result == '1':
                        is_find = 1
                        temp = get_word(words)
                        final_citation.append(temp)
                        current_title = title
                        final_table.append(table)
                elif is_find == 1:
                    if title == None:
                        temp = get_word(words)
                        final_citation.append(temp)
                        final_table.append(table)
                    else:
                        if result == '1':
                            output_table = merge_tables(final_table)
                            citation = read_citation(merge_word(final_citation))
                            write_to_excel(output_table, pmid, current_title, citation)
                            final_table = []
                            final_citation = []
                            temp = get_word(words)
                            final_citation.append(temp)
                            final_table.append(table)
                            current_title = title
                        else:
                            is_find = 0
                            output_table = merge_tables(final_table)
                            citation = read_citation(merge_word(final_citation))
                            write_to_excel(output_table, pmid, current_title, citation)
                            final_table = []
                            final_citation = []

        if is_find == 1:
            is_find = 0
            citation = read_citation(merge_word(final_citation))
            output_table = merge_tables(final_table)
            write_to_excel(output_table, pmid, current_title, citation)
            final_table = []
            final_citation = []
            
def read_quality_assessment_table(base_dir: str):
    """
    Read the specified base directory and process each PDF file found in 'supp' subdirectories.

    Parameters:
        base_dir (str): The base directory to search for PDF files.
    """
    total = 0   #Total number of articles read
    count = 0   #Total number of articles that has quality assessment table
    for root, dirs, files in os.walk(base_dir):
        if "supp" in dirs:
            pmid = os.path.basename(root)
            supp_path = os.path.join(root, "supp")
            print(pmid, supp_path)
            pdf_files = [f for f in os.listdir(supp_path) if f.lower().endswith(".pdf")]
            for pdf in pdf_files:
                pdf_path = os.path.join(supp_path, pdf)
                print(f"Quality Assessment\nroot: {root}, pmid: {pmid}, pdf: {pdf}")
                total += 1
                caption = read_caption(pdf_path)
                if has_quality_assessment(caption) == '1':
                    count += 1
                    extract_quality_assessment_table(pdf_path, pmid)
                    output_path = os.path.join(base_dir, 'selected_articles')
                    if not os.path.exists(output_path):  
                        os.makedirs(output_path)
                    if not os.path.exists(os.path.join(output_path, pdf)):
                        shutil.copyfile(pdf_path, os.path.join(output_path, pdf))
                    
    return total, count

# Example usage
#url = "data/jama_eye_disease/29459947/supp/jamainternmed-178-502-s001.pdf"
#extract_quality_assessment_table(url, "29459947", "data/jama_eye_disease/29459947/supp", table_type="Quality Assessment")
#read_quality_assessment_table("./data")
